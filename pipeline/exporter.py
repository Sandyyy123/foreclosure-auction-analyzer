"""Export filtered + ranked results to Excel with conditional formatting."""
import pandas as pd
from pathlib import Path

COLUMN_ORDER = [
    "case_number", "address", "appraised_value",
    "zestimate", "rent_zestimate", "price_to_rent_ratio",
    "year_built", "sqft", "bathrooms",
    "lien_flags", "warnings", "recommendation"
]

REC_COLORS = {"BUY": "C6EFCE", "HOLD": "FFEB9C", "SKIP": "FFC7CE", "Review": "D9D9D9"}

def export_ranked(df: pd.DataFrame, output_path: str) -> str:
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    # Sort by P/R ratio ascending (lower = better deal)
    df_sorted = df.sort_values("price_to_rent_ratio", na_position="last").reset_index(drop=True)

    cols = [c for c in COLUMN_ORDER if c in df_sorted.columns]
    df_out = df_sorted[cols]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Auction Analysis")
        ws = writer.sheets["Auction Analysis"]

        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # Color recommendation column
        from openpyxl.styles import PatternFill
        rec_col_idx = cols.index("recommendation") + 1 if "recommendation" in cols else None
        if rec_col_idx:
            for row in ws.iter_rows(min_row=2, min_col=rec_col_idx, max_col=rec_col_idx):
                for cell in row:
                    color = REC_COLORS.get(str(cell.value), "D9D9D9")
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    return output_path
